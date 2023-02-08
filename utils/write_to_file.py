import csv
import os
import openpyxl


def write_to_csv(filename, values):
    # filename is the name of the granular report
    # headers is an array containing the column names -
    # values is a 2D Array with each element signifying a particular row
    cwd = os.getcwd()
    file_path = f'{cwd}/{filename}.csv'

    with open(file_path, "a") as value:
        report = csv.writer(value)
        for row in values:
            report.writerow(row)


def create_new_excel_for_file(location, first_file, second_file):
    wb = openpyxl.Workbook()
    # Delete the default sheet
    wb.remove(wb.active)
    # Create new Sheet
    wb.create_sheet('scan-status')
    wb.create_sheet('source-&-sink-report')
    wb.create_sheet('flow-report')
    wb.save(location)


def create_new_excel(location, base_branch_name, head_branch_name):
    wb = openpyxl.Workbook()
    # Delete the default sheet
    wb.remove(wb.active)
    # Create new Sheet
    wb.create_sheet('scan-status')
    wb.create_sheet(f'{head_branch_name}-{base_branch_name}-source-&-sink-report')
    wb.create_sheet(f'{head_branch_name}-{base_branch_name}-flow-report')
    wb.create_sheet(f'{head_branch_name}-{base_branch_name}-performance-report')
    wb.save(location)


def write_source_sink_data(workbook_location, worksheet_name, report):
    # load the workbook
    workbook = openpyxl.load_workbook(filename=workbook_location)

    # use source-sink-collection sheet
    worksheet = workbook[worksheet_name]

    for row in report:
        worksheet.append(row)

    workbook.save(workbook_location)


def write_path_data(workbook_location, worksheet_name, report):
    workbook = openpyxl.load_workbook(filename=workbook_location)

    # create new sheet for source-sink-collection sheet
    worksheet = workbook[worksheet_name]

    for row in report:
        worksheet.append(row)

    workbook.save(workbook_location)


def write_performance_data(workbook_location, worksheet_name, report):
    workbook = openpyxl.load_workbook(filename=workbook_location)

    # create new sheet for source-sink-collection sheet
    worksheet = workbook[worksheet_name]

    for row in report:
        worksheet.append(row)

    workbook.save(workbook_location)


def write_scan_status_report(workbook_location, base_branch_name, head_branch_name, report):
    workbook = openpyxl.load_workbook(filename=workbook_location)

    # create new sheet for source-sink-collection sheet
    worksheet = workbook['scan-status']

    worksheet.append(["Repo", "Branch", "scan status", "scan error", "comparison status", "comparison error", "unique flow count", "scan time", "CPG size"])

    for repo in report.keys():
        repo_info = report[repo]
        worksheet.append([repo, base_branch_name, repo_info[base_branch_name]['scan_status'],
                          repo_info[base_branch_name]['scan_error_message'],
                          repo_info[base_branch_name]['comparison_status'],
                          repo_info[base_branch_name]['comparison_error_message'],
                          repo_info[base_branch_name]['unique_flows'],
                          repo_info[base_branch_name]['code_scan_time'],
                          repo_info[base_branch_name]['binary_file_size']])

        worksheet.append([repo, base_branch_name, repo_info[head_branch_name]['scan_status'],
                          repo_info[head_branch_name]['scan_error_message'],
                          repo_info[head_branch_name]['comparison_status'],
                          repo_info[head_branch_name]['comparison_error_message'],
                          repo_info[head_branch_name]['unique_flows'],
                          repo_info[head_branch_name]['code_scan_time'],
                          repo_info[head_branch_name]['binary_file_size']])

    workbook.save(workbook_location)